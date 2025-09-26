import streamlit as st
import pandas as pd
import tempfile
import zipfile
import io
import pyperclip  # For clipboard copy (optional, depends on backend environment)
import openpyxl
import os
import re
# Replace this with your actual backend function

def run_backend_on_files(extracted_path):
    print("Extracted path is",extracted_path)
    #For now hardcoded:
    start_row = 3
    end_row = 20
    start_column = 1
    end_column = 15
    pilos_dict=[]
    # for p in extracted_path:
    for root, dirs, files in os.walk(extracted_path):
        for f in files:
            p = os.path.join(root, f)
            # To open the workbook
            # workbook object is created
            wb_obj = openpyxl.load_workbook(p, data_only=True)

            # Get workbook active sheet object
            # from the active attribute
            for i in wb_obj.sheetnames:
                # print(i)
                if i =="PILOsReport":
                    sheet_obj=wb_obj[i]
                    data =[]
                    for row in sheet_obj.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
                        row_data = []
                        for cell in row:
                            row_data.append(cell.value)
                        data.append(row_data)

                    df_temp=pd.DataFrame(data)
                    instructor=df_temp.iloc[0]
                    pilos_asses=df_temp.iloc[17]
                    pilos_list=pilos_asses.values.tolist()
                    d={}
                    columns_list=["Subject", "PILOS","SO1","SO2","SO3","SO4","SO5","SO6","SO7"]
                    for j in columns_list:
                        # for i in pilos_list:
                        # print("I is at",i)
                        if j =="Subject":
                            subject = [splits.replace(".xlsx", "") for splits in p.split('\\') if "EENG" in splits][0]
                            d.update({j:subject})
                            print(j,p)
                        else:
                            d.update({j:pilos_list[columns_list.index(j)]})
                            print(j,pilos_list[columns_list.index(j)])
                    # df_master=df_master.append(pilos_dict,ignore_index=True)
                    pilos_dict.append(d)
                # print(d)

    pilos_df= pd.DataFrame(pilos_dict)

    final_pilos_df=calculate_average_pilos(pilos_df)
    return final_pilos_df

def calculate_average_pilos(df):
    so_columns = [col for col in df.columns if re.match(r'SO\d+', col)]
    averages=[]

    for col in df.columns:
        if col not in so_columns:
            averages.append("")
        else:
            list_of_percentages=df[col].dropna()
            averages.append(list_of_percentages.mean())
            # print("col",col, "list:",list_of_percentages)
    df.loc[len(df)] = averages

    return df


def get_students_info(extracted_path):
    st.text("Student Information")
    student_info=[]
    for root, dirs, files in os.walk(extracted_path):
        for f in files:
            p = os.path.join(root, f)
            # To open the workbook
            # workbook object is created
            wb_obj = openpyxl.load_workbook(p, data_only=True)
            data ={}
            for i in wb_obj.sheetnames:
                # print(i)
                
                if i =="Grades":
                    sheet_obj=wb_obj[i]
                    num_students = sheet_obj.cell(row=31, column=2).value
                    section = sheet_obj.cell(row=5, column=8).value
                    subject=sheet_obj.cell(row=2, column=8).value
                    data.update({ "Subject": subject,"Number of Students": num_students, "Section": section})
                    print("subject",subject)
                if i =="Marks":
                    row_idx = find_average_row(wb_obj[i])
                    sheet_obj=wb_obj[i]
                    avg = sheet_obj.cell(row=row_idx[0], column=88).value
                    data.update({"Average": avg})
                    print("data is",data)
                
            student_info.append(data)
    print("student info is",student_info)     
    student_df=pd.DataFrame(student_info)   
    return student_df



def find_average_row(sheet_obj):
    data = sheet_obj.values
    df = pd.DataFrame(data)
    row_idx = df[df.apply(lambda row: row.astype(str).str.contains("Average", case=False).any(), axis=1)].index
    return row_idx.tolist()  # list of row numbers where 'Average' appears



def student_tool(tmpdir):
    st.text("Loading each Subject's average grade, number of students, and section")
    df = get_students_info(tmpdir)
    # File uploader

    st.success("Processing complete ✅")

    # Display the dataframe
    st.dataframe(df)

    # Download option
    csv = df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "Download as CSV",
        csv,
        "results.csv",
        "text/csv",
        key="download-csv"
    )

    # Copy to clipboard option (browser limitation workaround)
    # Streamlit doesn’t allow direct clipboard copy on server,
    # so we use JS trick:
    st.text_area("Copy results:", df.to_csv(index=False), height=200)
    st.caption("Select all (Ctrl+A) and copy (Ctrl+C)")

    # Optional: If running locally and want automatic copy:
    if st.button("Copy to Clipboard (local only)"):
        try:
            pyperclip.copy(df.to_csv(index=False))
            st.success("Copied to clipboard!")
        except Exception as e:
            st.error(f"Clipboard copy failed: {e}")

def pilos(tmpdir):
    
    st.text("Loading the PILOs' Average Calculator")
    df = run_backend_on_files(tmpdir)
    # File uploader

    st.success("Processing complete ✅")

    # Display the dataframe
    st.dataframe(df)

    # Download option
    csv = df.to_csv(index=False).encode("utf-8")
    st.download_button(
        "Download as CSV",
        csv,
        "results.csv",
        "text/csv",
        key="download-csv"
    )

    # Copy to clipboard option (browser limitation workaround)
    # Streamlit doesn’t allow direct clipboard copy on server,
    # so we use JS trick:
    st.text_area("Copy results:", df.to_csv(index=False), height=200)
    st.caption("Select all (Ctrl+A) and copy (Ctrl+C)")

    # Optional: If running locally and want automatic copy:
    if st.button("Copy to Clipboard (local only)"):
        try:
            pyperclip.copy(df.to_csv(index=False))
            st.success("Copied to clipboard!")
        except Exception as e:
            st.error(f"Clipboard copy failed: {e}")



# ----------------- Streamlit App -----------------
st.title("QAC ToolBox")

uploaded_file = st.file_uploader("Upload a ZIP file", type=["zip"])

if uploaded_file is not None:
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(uploaded_file, "r") as z:
            z.extractall(tmpdir)

        page_names_to_funcs = {
            "PiLOs": pilos,
            "Students": student_tool,
        }

        demo_name = st.sidebar.selectbox("Choose a tool", page_names_to_funcs.keys())
        page_names_to_funcs[demo_name](tmpdir)